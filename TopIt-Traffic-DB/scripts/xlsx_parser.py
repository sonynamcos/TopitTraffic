"""
주기표 엑셀 파서 - .xls/.xlsx 파일에서 교차로별 주기표 정보를 추출한다.

엑셀 파일 구조:
  - 범위 파일: "0~10.xls" → 시트마다 교차로 1개 (시트명 = 교차로명)
  - 개별 파일: "해날아파트주기표.xlsx" → 단일 교차로
  - 신규 파일: "01_20 신규.xls" → 시트별 교차로
"""

import os
import re
from pathlib import Path
from typing import Optional

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import xlrd
except ImportError:
    xlrd = None


# 무시할 시트명 (정확 일치, 소문자)
SKIP_SHEET_NAMES = {
    'sheet1', 'sheet2', 'sheet3', 'sheet',
    '요약', '목록', '메모', '변경내용', '참고',
    '표지', 'cover', 'index', 'summary', 'test', 'ts',
    '설계', '설계내역', '설계서표지', '총괄내역', '집계내역표',
    '원가계산', '공정표', '도면', '내역서', '노임',
    '샘플', '연동', '공양식',
}

# 무시할 시트명 패턴 (부분 일치)
SKIP_SHEET_PATTERNS = [
    r'^sheet\d*$',
    r'내역', r'원가', r'공정', r'설계', r'도면', r'표지',
    r'집계', r'총괄', r'샘플', r'연동화',
]


def parse_excel_file(filepath: str) -> list[dict]:
    """엑셀 파일을 열어 시트별 교차로 정보를 추출한다.

    Returns:
        시트별 교차로 정보 딕셔너리 리스트:
        [{
            'source_file': str,
            'sheet_name': str,
            'intersection_name': str | None,
            'intersection_number': int | None,
            'has_content': bool,
        }]
    """
    filepath = str(filepath)
    ext = os.path.splitext(filepath)[1].lower()

    if ext == '.xlsx':
        return _parse_xlsx(filepath)
    elif ext == '.xls':
        return _parse_xls(filepath)
    else:
        return []


def _parse_xlsx(filepath: str) -> list[dict]:
    if openpyxl is None:
        return [_error_entry(filepath, 'openpyxl 미설치')]

    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        return [_error_entry(filepath, str(e))]

    results = []
    for sheet_name in wb.sheetnames:
        entry = _make_entry(filepath, sheet_name)

        ws = wb[sheet_name]
        # 시트 내용 확인 (처음 10행만)
        row_count = 0
        cell_texts = []
        try:
            for row in ws.iter_rows(max_row=10, values_only=True):
                row_count += 1
                for cell in row:
                    if cell is not None:
                        cell_texts.append(str(cell).strip())
        except Exception:
            pass

        entry['has_content'] = row_count > 0 and len(cell_texts) > 0

        # 시트명에서 교차로명 추출
        name = _extract_intersection_name(sheet_name)
        if name:
            entry['intersection_name'] = name
        elif cell_texts:
            # 시트명이 범용이면 셀에서 교차로명 추출 시도
            name = _extract_name_from_cells(cell_texts)
            if name:
                entry['intersection_name'] = name

        results.append(entry)

    wb.close()
    return results


def _parse_xls(filepath: str) -> list[dict]:
    if xlrd is None:
        return [_error_entry(filepath, 'xlrd 미설치')]

    try:
        wb = xlrd.open_workbook(filepath, on_demand=True)
    except Exception as e:
        return [_error_entry(filepath, str(e))]

    results = []
    for sheet_name in wb.sheet_names():
        entry = _make_entry(filepath, sheet_name)

        try:
            ws = wb.sheet_by_name(sheet_name)
            entry['has_content'] = ws.nrows > 0 and ws.ncols > 0

            # 교차로명 추출
            name = _extract_intersection_name(sheet_name)
            if name:
                entry['intersection_name'] = name
            elif ws.nrows > 0:
                cell_texts = []
                for r in range(min(ws.nrows, 10)):
                    for c in range(min(ws.ncols, 10)):
                        try:
                            val = ws.cell_value(r, c)
                            if val:
                                cell_texts.append(str(val).strip())
                        except Exception:
                            pass
                name = _extract_name_from_cells(cell_texts)
                if name:
                    entry['intersection_name'] = name
        except Exception:
            entry['has_content'] = False

        results.append(entry)

    return results


def _make_entry(filepath: str, sheet_name: str) -> dict:
    return {
        'source_file': filepath,
        'source_filename': os.path.basename(filepath),
        'sheet_name': sheet_name,
        'intersection_name': None,
        'intersection_number': None,
        'has_content': False,
        'error': None,
    }


def _error_entry(filepath: str, error: str) -> dict:
    entry = _make_entry(filepath, '')
    entry['error'] = error
    return entry


def _extract_intersection_name(sheet_name: str) -> Optional[str]:
    """시트명에서 교차로명을 추출한다."""
    name = sheet_name.strip()

    name_lower = name.lower()

    # 범용/비교차로 시트명 스킵
    if name_lower in SKIP_SHEET_NAMES:
        return None

    # 패턴 기반 스킵
    for pat in SKIP_SHEET_PATTERNS:
        if re.search(pat, name_lower):
            return None

    # 숫자만 있는 시트명 스킵
    if re.match(r'^\d+$', name):
        return None

    # 너무 짧은 시트명 (1글자) 스킵
    if len(name.strip()) <= 1:
        return None

    # 앞쪽 숫자 + 언더스코어/점 제거: "047_신설사거리" → "신설사거리"
    num_match = re.match(r'^(\d+)[_.\-\s]+(.+)$', name)
    cleaned = name
    if num_match:
        cleaned = num_match.group(2)

    # "주기표" 접미사 제거
    cleaned = re.sub(r'[_\s]?주기표$', '', cleaned)
    # 날짜 접미사 제거 (다양한 형식)
    cleaned = re.sub(r'[\._\- ]?\d{4}\.\d{2}\.\d{2}\.?$', '', cleaned)
    cleaned = re.sub(r'[\._\- ]?\d{2}\.\d{2}\.\d{2}\.?$', '', cleaned)
    cleaned = re.sub(r'[\._\- ]?\d{8}$', '', cleaned)
    cleaned = re.sub(r'[\._\- ]?\d{6}$', '', cleaned)
    # 모든 괄호 내용 제거
    cleaned = re.sub(r'\([^)]*\)', '', cleaned)
    cleaned = re.sub(r'\($', '', cleaned)
    # old/new/변경 등 접미사 제거
    for _ in range(3):
        cleaned = re.sub(
            r'[_\-,\s]?(old|new|구|신|변경|수정|기존|최신|copy|백업|임시)$',
            '', cleaned, flags=re.I
        )
    # 숫자 접미사: "4R", "3R"
    cleaned = re.sub(r'\d+[Rr]$', '', cleaned)
    # 제조사 접미사
    cleaned = re.sub(r'[_\- ]?(서돌|한진|서|한)$', '', cleaned)
    # 후행 숫자 제거
    cleaned = re.sub(r'\d+$', '', cleaned)
    # 앞뒤 공백/특수문자 정리
    cleaned = cleaned.strip(' _-.,')

    if not cleaned:
        return None

    return cleaned


def _extract_name_from_cells(cell_texts: list[str]) -> Optional[str]:
    """셀 내용에서 교차로명을 추출 시도한다."""
    # "교차로명", "교차로", "위치" 등의 레이블 다음 셀 값
    keywords = ['교차로명', '교차로', '위치', '지점명', '교차로 명']

    for i, text in enumerate(cell_texts):
        for kw in keywords:
            if kw in text:
                # 같은 텍스트에서 콜론 뒤
                colon_match = re.search(r'[:：]\s*(.+)', text)
                if colon_match:
                    return colon_match.group(1).strip()
                # 다음 셀
                if i + 1 < len(cell_texts):
                    next_val = cell_texts[i + 1]
                    if next_val and not any(k in next_val for k in keywords):
                        return next_val.strip()

    # "~사거리", "~삼거리", "~교차로" 패턴 검색
    for text in cell_texts:
        if re.search(r'(사거리|삼거리|교차로|입구|정문|후문|앞|초등학교|중학교|고등학교)', text):
            # 너무 긴 텍스트는 교차로명이 아닐 가능성
            if len(text) <= 20:
                return text.strip()

    return None


def scan_excel_directory(directory: str) -> list[dict]:
    """디렉토리 내 모든 엑셀 파일을 스캔하여 시트별 교차로 정보를 반환한다."""
    results = []
    dir_path = Path(directory)

    for ext_pattern in ['*.xlsx', '*.xls']:
        for excel_file in sorted(dir_path.rglob(ext_pattern)):
            # 임시 파일 스킵
            if excel_file.name.startswith('~$'):
                continue
            # 요도 파일 스킵
            if '요도' in excel_file.name:
                continue
            # DNG/이미지 파일 스킵
            if excel_file.suffix.lower() in ('.dng', '.png', '.jpg'):
                continue
            # 설계/도면 관련 파일 스킵
            skip_keywords = ['설계', '도면', 'DBSheet', '제어기조사']
            if any(kw in excel_file.name for kw in skip_keywords):
                continue
            # zip 파일 스킵
            if excel_file.suffix.lower() == '.zip':
                continue

            try:
                sheets = parse_excel_file(str(excel_file))
                results.extend(sheets)
            except Exception as e:
                results.append(_error_entry(str(excel_file), str(e)))

    return results


def extract_cycle_table_data(filepath: str, sheet_name: str) -> Optional[dict]:
    """특정 시트에서 주기표 데이터(주기, 현시 등)를 상세 추출한다."""
    # 향후 확장: 주기표의 세부 타이밍 데이터 추출
    # 현재는 시트 존재 여부와 교차로명만 추출
    return None


if __name__ == '__main__':
    import json
    import sys

    target = sys.argv[1] if len(sys.argv) > 1 else '.'

    if os.path.isfile(target):
        results = parse_excel_file(target)
        for r in results:
            print(json.dumps(r, ensure_ascii=False, indent=2))
    else:
        results = scan_excel_directory(target)
        total = len(results)
        named = sum(1 for r in results if r['intersection_name'])
        errors = sum(1 for r in results if r.get('error'))
        print(f'총 {total}개 시트 스캔 (교차로명 추출: {named}, 오류: {errors})')
        for r in results:
            status = '✅' if r['intersection_name'] else '❌'
            name = r['intersection_name'] or '(추출실패)'
            sheet = r['sheet_name']
            src = r['source_filename']
            print(f'  {status} {src:40s} [{sheet:20s}] → {name}')
